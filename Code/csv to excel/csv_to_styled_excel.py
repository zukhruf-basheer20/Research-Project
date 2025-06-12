import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font

# === CONFIG ===
CSV_PATH = Path("../CSV Files/EffNetB0_leaf_deep_features_with_metadata.csv")
EXCEL_PATH = CSV_PATH.with_suffix(".xlsx")

# === LOAD CSV ===
print(f"📄 Loading CSV from: {CSV_PATH}")
df = pd.read_csv(CSV_PATH)

# === SAVE TO EXCEL ===
print("📊 Converting to Excel...")
df.to_excel(EXCEL_PATH, index=False)

# === STYLE HEADER ===
wb = load_workbook(EXCEL_PATH)
ws = wb.active

for cell in ws[1]:
    cell.font = Font(bold=True, size=14)

wb.save(EXCEL_PATH)
print(f"✅ Styled Excel file saved at: {EXCEL_PATH}")
